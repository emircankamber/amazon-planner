"""
main.py — Amazon Stok Planlama (multi-user, session auth)
"""
from __future__ import annotations

import json
import math
import statistics
from datetime import date
from io import BytesIO
from typing import List, Optional, Tuple

from fastapi import Cookie, FastAPI, Form, HTTPException, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from auth import (
    authenticate_user,
    create_user,
    decode_session_cookie,
    get_current_user,
    make_session_cookie,
)
from db import get_conn, init_db

app = FastAPI(title="Amazon Stok Planlama")
init_db()


# ══════════════════════════════════════════════════════════════════════════════
#  Helpers — tarih / hesap
# ══════════════════════════════════════════════════════════════════════════════

def last_n_calendar_months(n: int) -> List[Tuple[int, int]]:
    today = date.today()
    y, m = today.year, today.month
    out: List[Tuple[int, int]] = []
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return out


def month_label(y: int, m: int) -> str:
    return f"{y}-{m:02d}"


def compute_from_last_months(
    lead_time: int,
    z: float,
    monthly_units: List[int],
    fba_stock: int,
    inbound_stock: int,
) -> dict:
    if not monthly_units:
        return {
            "daily_velocity": 0.0,
            "std_daily": 0.0,
            "safety_stock": 0.0,
            "rop": 0.0,
            "order_qty": 0.0,
        }
    mean_month = sum(monthly_units) / len(monthly_units)
    daily_velocity = mean_month / 30.0
    std_daily = statistics.stdev(monthly_units) / 30.0 if len(monthly_units) >= 2 else 0.0
    safety_stock = z * std_daily * math.sqrt(max(1, lead_time))
    rop = daily_velocity * lead_time + safety_stock
    order_qty = max(
        0.0,
        daily_velocity * 60 + safety_stock - (fba_stock + inbound_stock),
    )
    return {
        "daily_velocity": daily_velocity,
        "std_daily": std_daily,
        "safety_stock": safety_stock,
        "rop": rop,
        "order_qty": order_qty,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  DB helpers (user-scoped)
# ══════════════════════════════════════════════════════════════════════════════

def upsert_product(cur, user_id: int, sku: str, name: str, lead_time_days: int,
                   z_value: float, fba_stock: int, inbound_stock: int) -> None:
    cur.execute("""
        INSERT INTO products(user_id, sku, name, lead_time_days, z_value, fba_stock, inbound_stock)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(user_id, sku) DO UPDATE SET
            name=excluded.name,
            lead_time_days=excluded.lead_time_days,
            z_value=excluded.z_value,
            fba_stock=excluded.fba_stock,
            inbound_stock=excluded.inbound_stock,
            updated_at=CURRENT_TIMESTAMP;
    """, (user_id, sku, name, lead_time_days, z_value, fba_stock, inbound_stock))


def upsert_monthly_sales(cur, user_id: int, sku: str,
                         years: List[int], months: List[int], units_sold: List[int]) -> None:
    for y, m, u in zip(years, months, units_sold):
        cur.execute("""
            INSERT INTO monthly_sales(user_id, sku, year, month, units_sold)
            VALUES(?,?,?,?,?)
            ON CONFLICT(user_id, sku, year, month) DO UPDATE SET
                units_sold=excluded.units_sold,
                created_at=CURRENT_TIMESTAMP;
        """, (user_id, sku, int(y), int(m), int(u)))


def fetch_product(cur, user_id: int, sku: str):
    return cur.execute(
        "SELECT * FROM products WHERE user_id=? AND sku=?", (user_id, sku)
    ).fetchone()


def fetch_month_units(cur, user_id: int, sku: str, y: int, m: int) -> Optional[int]:
    row = cur.execute(
        "SELECT units_sold FROM monthly_sales WHERE user_id=? AND sku=? AND year=? AND month=?",
        (user_id, sku, y, m),
    ).fetchone()
    return None if row is None else int(row["units_sold"])


def delete_product_and_sales(cur, user_id: int, sku: str) -> None:
    cur.execute("DELETE FROM monthly_sales WHERE user_id=? AND sku=?", (user_id, sku))
    cur.execute("DELETE FROM products WHERE user_id=? AND sku=?", (user_id, sku))


def delete_single_month_sale(cur, user_id: int, sku: str, year: int, month: int) -> None:
    cur.execute(
        "DELETE FROM monthly_sales WHERE user_id=? AND sku=? AND year=? AND month=?",
        (user_id, sku, year, month),
    )


def compute_for_sku(cur, user_id: int, sku: str):
    prod = fetch_product(cur, user_id, sku)
    if prod is None:
        return None
    last3 = last_n_calendar_months(3)
    monthly_units: List[int] = []
    for y, m in last3:
        u = fetch_month_units(cur, user_id, sku, y, m)
        if u is not None:
            monthly_units.append(u)
    res = compute_from_last_months(
        lead_time=int(prod["lead_time_days"]),
        z=float(prod["z_value"]),
        monthly_units=monthly_units,
        fba_stock=int(prod["fba_stock"]),
        inbound_stock=int(prod["inbound_stock"]),
    )
    return prod, last3, monthly_units, res


# ══════════════════════════════════════════════════════════════════════════════
#  UI helpers
# ══════════════════════════════════════════════════════════════════════════════

def _nav(logged_in: bool = True) -> str:
    if not logged_in:
        return ""
    return """
      <div class="flex gap-2 flex-wrap">
        <a class="px-4 py-2 rounded-2xl text-sm border bg-slate-900/40 border-slate-700/60 hover:bg-slate-800/60 transition" href="/">Ekle</a>
        <a class="px-4 py-2 rounded-2xl text-sm border bg-slate-900/40 border-slate-700/60 hover:bg-slate-800/60 transition" href="/products">SKU Listesi</a>
        <a class="px-4 py-2 rounded-2xl text-sm border bg-slate-900/40 border-slate-700/60 hover:bg-slate-800/60 transition" href="/plan">Sipariş Planı</a>
        <form method="post" action="/logout">
          <button type="submit" class="px-4 py-2 rounded-2xl text-sm border border-rose-500/40 bg-rose-500/10 text-rose-300 hover:bg-rose-500/20 transition">
            Çıkış Yap
          </button>
        </form>
      </div>
    """


def page_shell(title: str, body_html: str, logged_in: bool = True) -> str:
    return f"""<!doctype html>
<html lang="tr">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <script src="https://cdn.tailwindcss.com"></script>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <title>{title} — Amazon Stok Planlama</title>
  <style>
    input:focus {{ outline: none; box-shadow: 0 0 0 2px #6366f1; }}
    .card {{ border-radius: 1.5rem; border: 1px solid rgba(100,116,139,.35); background: rgba(15,23,42,.45); padding: 1.5rem; box-shadow: 0 4px 24px #0004; }}
  </style>
</head>
<body class="min-h-screen bg-gradient-to-br from-slate-950 via-slate-900 to-slate-950 text-slate-100 antialiased">
  <div class="max-w-6xl mx-auto px-4 py-10">
    <div class="flex items-start justify-between gap-6 flex-wrap mb-8">
      <div>
        <h1 class="text-3xl md:text-4xl font-extrabold tracking-tight bg-gradient-to-r from-indigo-400 to-sky-400 bg-clip-text text-transparent">
          Amazon Stok Planlama
        </h1>
        <p class="text-slate-400 mt-1 text-sm">{title}</p>
      </div>
      {_nav(logged_in)}
    </div>
    {body_html}
  </div>
</body>
</html>"""


def _input(name: str, placeholder: str, type_: str = "text", value: str = "",
           extra: str = "", required: bool = True) -> str:
    req = "required" if required else ""
    return (f'<input name="{name}" type="{type_}" placeholder="{placeholder}" value="{value}" '
            f'class="w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3 '
            f'text-sm placeholder-slate-500 transition" {extra} {req}/>')


def build_default_rows_html() -> str:
    rows = []
    for y, m in last_n_calendar_months(3):
        rows.append(f"""
        <tr class="border-b border-slate-700/40">
          <td class="py-2 pr-1"><input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2 text-sm" name="years" type="number" value="{y}" required/></td>
          <td class="py-2 pr-1"><input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2 text-sm" name="months" type="number" min="1" max="12" value="{m}" required/></td>
          <td class="py-2 pr-1"><input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2 text-sm" name="units_sold" type="number" min="0" placeholder="Adet" required/></td>
          <td class="py-2 text-right">
            <button type="button" class="remove-row text-xs px-3 py-2 rounded-xl bg-rose-500/15 text-rose-300 border border-rose-500/30 hover:bg-rose-500/25 transition">Kaldır</button>
          </td>
        </tr>""")
    return "\n".join(rows)


# ══════════════════════════════════════════════════════════════════════════════
#  Auth — cookie helper
# ══════════════════════════════════════════════════════════════════════════════

def _current_user_id(session: Optional[str]) -> Optional[int]:
    if not session:
        return None
    return decode_session_cookie(session)


def _require_user(session: Optional[str]) -> int:
    uid = _current_user_id(session)
    if uid is None:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/login"},
        )
    return uid


# ══════════════════════════════════════════════════════════════════════════════
#  Auth Routes
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/login", response_class=HTMLResponse)
def login_page(error: str = "", session: Optional[str] = Cookie(default=None)):
    if _current_user_id(session):
        return RedirectResponse("/", status_code=302)
    err_html = f'<p class="text-rose-400 text-sm mt-3">{error}</p>' if error else ""
    body = f"""
<div class="flex justify-center">
  <div class="w-full max-w-md">
    <div class="card">
      <h2 class="text-2xl font-bold mb-6">Giriş Yap</h2>
      <form method="post" action="/login" class="space-y-4">
        {_input("email", "E-posta", "email")}
        {_input("password", "Şifre", "password")}
        {err_html}
        <button type="submit" class="w-full rounded-2xl bg-indigo-500 hover:bg-indigo-400 transition px-5 py-3 font-semibold text-sm mt-2">
          Giriş Yap
        </button>
      </form>
      <p class="text-slate-400 text-sm mt-5 text-center">
        Hesabın yok mu? <a class="text-indigo-400 hover:underline" href="/register">Kayıt ol</a>
      </p>
    </div>
  </div>
</div>"""
    return page_shell("Giriş", body, logged_in=False)


@app.post("/login")
def login(
    email: str = Form(...),
    password: str = Form(...),
):
    uid = authenticate_user(email, password)
    if uid is None:
        return RedirectResponse(
            url=f"/login?error=E-posta+veya+şifre+hatalı",
            status_code=303,
        )
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        "session",
        make_session_cookie(uid),
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 24 * 30,
        secure=False,  # Render HTTPS'te True yapılabilir
    )
    return response


@app.get("/register", response_class=HTMLResponse)
def register_page(error: str = "", session: Optional[str] = Cookie(default=None)):
    if _current_user_id(session):
        return RedirectResponse("/", status_code=302)
    err_html = f'<p class="text-rose-400 text-sm mt-3">{error}</p>' if error else ""
    body = f"""
<div class="flex justify-center">
  <div class="w-full max-w-md">
    <div class="card">
      <h2 class="text-2xl font-bold mb-6">Kayıt Ol</h2>
      <form method="post" action="/register" class="space-y-4">
        {_input("email", "E-posta", "email")}
        {_input("password", "Şifre (en az 8 karakter)", "password")}
        {_input("password2", "Şifre tekrar", "password")}
        {err_html}
        <button type="submit" class="w-full rounded-2xl bg-indigo-500 hover:bg-indigo-400 transition px-5 py-3 font-semibold text-sm mt-2">
          Hesap Oluştur
        </button>
      </form>
      <p class="text-slate-400 text-sm mt-5 text-center">
        Zaten hesabın var mı? <a class="text-indigo-400 hover:underline" href="/login">Giriş yap</a>
      </p>
    </div>
  </div>
</div>"""
    return page_shell("Kayıt Ol", body, logged_in=False)


@app.post("/register")
def register(
    email: str = Form(...),
    password: str = Form(...),
    password2: str = Form(...),
):
    if len(password) < 8:
        return RedirectResponse("/register?error=Şifre+en+az+8+karakter+olmalı", status_code=303)
    if password != password2:
        return RedirectResponse("/register?error=Şifreler+eşleşmiyor", status_code=303)
    try:
        uid = create_user(email, password)
    except ValueError as exc:
        import urllib.parse
        return RedirectResponse(
            f"/register?error={urllib.parse.quote(str(exc))}", status_code=303
        )
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        "session",
        make_session_cookie(uid),
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 24 * 30,
        secure=False,
    )
    return response


@app.post("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("session")
    return response


# ══════════════════════════════════════════════════════════════════════════════
#  App Routes
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
def home(session: Optional[str] = Cookie(default=None)):
    try:
        _require_user(session)
    except HTTPException:
        return RedirectResponse("/login", status_code=302)

    body = f"""
<form method="post" action="/upsert" class="grid grid-cols-1 lg:grid-cols-3 gap-6">
  <!-- Ürün bilgileri -->
  <div class="card lg:col-span-1">
    <h2 class="text-lg font-semibold mb-4">Ürün Bilgileri</h2>
    <div class="space-y-3">
      {_input("sku", "SKU *")}
      {_input("name", "Ürün adı (opsiyonel)", required=False)}
      <div class="grid grid-cols-2 gap-3">
        {_input("lead_time_days", "Lead time (gün)", "number", extra="min='1'")}
        {_input("z_value", "Z değeri (örn: 1.65)", "number", extra="step='0.01' min='0'")}
      </div>
      <div class="grid grid-cols-2 gap-3">
        {_input("fba_stock", "FBA stok", "number", value="0", extra="min='0'")}
        {_input("inbound_stock", "Yoldaki stok", "number", value="0", extra="min='0'")}
      </div>
    </div>
    <button type="submit" class="mt-6 w-full rounded-2xl bg-indigo-500 hover:bg-indigo-400 transition px-5 py-3 font-semibold text-sm">
      Kaydet &amp; Hesapla
    </button>
  </div>

  <!-- Aylık satışlar -->
  <div class="card lg:col-span-2">
    <div class="flex items-center justify-between gap-3 flex-wrap mb-4">
      <h2 class="text-lg font-semibold">Aylık Satışlar</h2>
      <button type="button" id="addRow"
        class="rounded-2xl bg-emerald-500/15 text-emerald-300 border border-emerald-500/30 hover:bg-emerald-500/25 transition px-4 py-2 text-sm font-semibold">
        + Ay Ekle
      </button>
    </div>
    <div class="overflow-x-auto">
      <table class="w-full text-sm">
        <thead class="text-slate-400">
          <tr class="border-b border-slate-700/60">
            <th class="text-left py-2 pr-2">Yıl</th>
            <th class="text-left py-2 pr-2">Ay</th>
            <th class="text-left py-2 pr-2">Satış Adedi</th>
            <th class="text-right py-2"></th>
          </tr>
        </thead>
        <tbody id="salesBody">
          {build_default_rows_html()}
        </tbody>
      </table>
    </div>
    <p class="text-slate-500 text-xs mt-3">Aynı (yıl, ay) girilirse mevcut kayıt güncellenir.</p>
  </div>
</form>

<script>
  const salesBody = document.getElementById("salesBody");
  function attachRemove() {{
    document.querySelectorAll(".remove-row").forEach(btn => {{
      btn.onclick = e => {{
        if (salesBody.querySelectorAll("tr").length > 1)
          e.target.closest("tr").remove();
      }};
    }});
  }}
  document.getElementById("addRow").addEventListener("click", () => {{
    const now = new Date();
    const tr = document.createElement("tr");
    tr.className = "border-b border-slate-700/40";
    tr.innerHTML = `
      <td class="py-2 pr-1"><input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2 text-sm" name="years" type="number" value="${{now.getFullYear()}}" required/></td>
      <td class="py-2 pr-1"><input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2 text-sm" name="months" type="number" min="1" max="12" value="${{now.getMonth()+1}}" required/></td>
      <td class="py-2 pr-1"><input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2 text-sm" name="units_sold" type="number" min="0" placeholder="Adet" required/></td>
      <td class="py-2 text-right"><button type="button" class="remove-row text-xs px-3 py-2 rounded-xl bg-rose-500/15 text-rose-300 border border-rose-500/30 hover:bg-rose-500/25 transition">Kaldır</button></td>`;
    salesBody.prepend(tr);
    attachRemove();
  }});
  attachRemove();
</script>"""
    return page_shell("Yeni Ürün / Satış Ekle", body)


@app.post("/upsert")
def upsert(
    session: Optional[str] = Cookie(default=None),
    sku: str = Form(...),
    name: str = Form(""),
    lead_time_days: int = Form(...),
    z_value: float = Form(...),
    fba_stock: int = Form(0),
    inbound_stock: int = Form(0),
    years: List[int] = Form(...),
    months: List[int] = Form(...),
    units_sold: List[int] = Form(...),
):
    user_id = _require_user(session)
    sku = sku.strip().upper()
    if not sku:
        return RedirectResponse("/?error=SKU+boş+olamaz", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    upsert_product(cur, user_id, sku, name.strip(), lead_time_days, z_value, fba_stock, inbound_stock)
    upsert_monthly_sales(cur, user_id, sku, years, months, units_sold)
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/product/{sku}", status_code=303)


@app.get("/products", response_class=HTMLResponse)
def products(session: Optional[str] = Cookie(default=None)):
    user_id = _get_user_or_redirect(session)
    if isinstance(user_id, RedirectResponse):
        return user_id

    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute(
        "SELECT sku FROM products WHERE user_id=? ORDER BY updated_at DESC", (user_id,)
    ).fetchall()]

    rows = []
    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, user_id, sku)
        oq = int(round(res["order_qty"]))
        badge = ("bg-emerald-500/15 text-emerald-300 border-emerald-500/30"
                 if oq <= 0 else
                 "bg-rose-500/15 text-rose-300 border-rose-500/30")
        rows.append(f"""
        <tr class="border-b border-slate-700/40 hover:bg-slate-800/20 transition">
          <td class="py-3 pr-3 font-semibold">{sku}</td>
          <td class="py-3 pr-3 text-slate-300 text-sm">{prod["name"] or "—"}</td>
          <td class="py-3 pr-3 text-right text-sm">{res["daily_velocity"]:.2f}</td>
          <td class="py-3 pr-3 text-right text-sm">{res["rop"]:.2f}</td>
          <td class="py-3 pr-3 text-right">
            <span class="px-3 py-1 rounded-2xl border text-xs font-semibold {badge}">{oq}</span>
          </td>
          <td class="py-3 text-right">
            <div class="flex justify-end gap-2">
              <a class="px-3 py-1.5 rounded-xl text-xs border border-slate-700/60 bg-slate-900/40 hover:bg-slate-800/60 transition" href="/product/{sku}">Detay</a>
              <form method="post" action="/delete/product/{sku}" onsubmit="return confirm('{sku} ve TÜM satış verileri silinsin mi?');">
                <button type="submit" class="px-3 py-1.5 rounded-xl text-xs border border-rose-500/30 bg-rose-500/10 text-rose-300 hover:bg-rose-500/20 transition">Sil</button>
              </form>
            </div>
          </td>
        </tr>""")
    conn.close()

    body = f"""
<div class="card">
  <div class="flex items-center justify-between flex-wrap gap-3 mb-5">
    <h2 class="text-xl font-semibold">SKU Listesi</h2>
    <div class="flex gap-2">
      <a href="/export/products.xlsx" class="px-4 py-2 rounded-2xl text-xs border border-slate-700/60 bg-slate-900/40 hover:bg-slate-800/60 transition font-semibold">⬇ SKU Excel</a>
      <a href="/export/plan.xlsx" class="px-4 py-2 rounded-2xl text-xs border border-slate-700/60 bg-slate-900/40 hover:bg-slate-800/60 transition font-semibold">⬇ Plan Excel</a>
    </div>
  </div>
  <div class="overflow-x-auto">
    <table class="w-full text-sm">
      <thead class="text-slate-400 text-xs uppercase tracking-wide">
        <tr class="border-b border-slate-700/60">
          <th class="text-left py-2 pr-3">SKU</th>
          <th class="text-left py-2 pr-3">Ürün</th>
          <th class="text-right py-2 pr-3">Günlük Hız</th>
          <th class="text-right py-2 pr-3">ROP</th>
          <th class="text-right py-2 pr-3">Sipariş</th>
          <th class="text-right py-2"></th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows) if rows else '<tr><td class="py-6 text-slate-500 text-center" colspan="6">Henüz ürün yok. <a class="text-indigo-400 underline" href="/">Ekle</a></td></tr>'}
      </tbody>
    </table>
  </div>
</div>"""
    return page_shell("SKU Listesi", body)


@app.get("/plan", response_class=HTMLResponse)
def plan(session: Optional[str] = Cookie(default=None)):
    user_id = _get_user_or_redirect(session)
    if isinstance(user_id, RedirectResponse):
        return user_id

    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute(
        "SELECT sku FROM products WHERE user_id=? ORDER BY updated_at DESC", (user_id,)
    ).fetchall()]

    rows = []
    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, user_id, sku)
        oq = int(round(res["order_qty"]))
        if oq <= 0:
            continue
        rows.append(f"""
        <tr class="border-b border-slate-700/40 hover:bg-slate-800/20 transition">
          <td class="py-3 pr-3 font-semibold">{sku}</td>
          <td class="py-3 pr-3 text-slate-300 text-sm">{prod["name"] or "—"}</td>
          <td class="py-3 pr-3 text-right text-sm">{int(prod["fba_stock"])}</td>
          <td class="py-3 pr-3 text-right text-sm">{int(prod["inbound_stock"])}</td>
          <td class="py-3 pr-3 text-right text-sm">{res["rop"]:.2f}</td>
          <td class="py-3 pr-3 text-right">
            <span class="px-3 py-1 rounded-2xl border text-xs font-semibold bg-rose-500/15 text-rose-300 border-rose-500/30">{oq}</span>
          </td>
          <td class="py-3 text-right">
            <a class="px-3 py-1.5 rounded-xl text-xs border border-slate-700/60 bg-slate-900/40 hover:bg-slate-800/60 transition" href="/product/{sku}">Detay</a>
          </td>
        </tr>""")
    conn.close()

    body = f"""
<div class="card">
  <div class="flex items-center justify-between flex-wrap gap-3 mb-5">
    <div>
      <h2 class="text-xl font-semibold">Sipariş Planı</h2>
      <p class="text-slate-400 text-xs mt-1">Yalnızca sipariş vermesi gereken ürünler listelenir.</p>
    </div>
    <a href="/export/plan.xlsx" class="px-4 py-2 rounded-2xl text-xs border border-slate-700/60 bg-slate-900/40 hover:bg-slate-800/60 transition font-semibold">⬇ Plan Excel</a>
  </div>
  <div class="overflow-x-auto">
    <table class="w-full text-sm">
      <thead class="text-slate-400 text-xs uppercase tracking-wide">
        <tr class="border-b border-slate-700/60">
          <th class="text-left py-2 pr-3">SKU</th>
          <th class="text-left py-2 pr-3">Ürün</th>
          <th class="text-right py-2 pr-3">FBA</th>
          <th class="text-right py-2 pr-3">Yoldaki</th>
          <th class="text-right py-2 pr-3">ROP</th>
          <th class="text-right py-2 pr-3">Sipariş (60g)</th>
          <th class="text-right py-2"></th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows) if rows else '<tr><td class="py-6 text-slate-500 text-center" colspan="7">Sipariş gereken ürün yok. 🎉</td></tr>'}
      </tbody>
    </table>
  </div>
</div>"""
    return page_shell("Sipariş Planı", body)


@app.get("/product/{sku}", response_class=HTMLResponse)
def product_detail(sku: str, session: Optional[str] = Cookie(default=None)):
    user_id = _get_user_or_redirect(session)
    if isinstance(user_id, RedirectResponse):
        return user_id

    conn = get_conn()
    cur = conn.cursor()
    computed = compute_for_sku(cur, user_id, sku)
    if computed is None:
        conn.close()
        return page_shell("Bulunamadı", '<div class="card text-slate-400">Bu SKU size ait değil veya mevcut değil.</div>')

    prod, last3, monthly_found, res = computed
    last6 = list(reversed(last_n_calendar_months(6)))
    labels6 = [month_label(y, m) for y, m in last6]
    units6 = [(fetch_month_units(cur, user_id, sku, y, m) or 0) for y, m in last6]
    conn.close()

    oq = int(round(res["order_qty"]))
    oq_color = "text-rose-300" if oq > 0 else "text-emerald-300"

    body = f"""
<div class="grid grid-cols-1 lg:grid-cols-3 gap-6">
  <!-- Sol panel -->
  <div class="card lg:col-span-1 space-y-4">
    <div>
      <div class="text-xs text-slate-400 uppercase tracking-wide">SKU</div>
      <div class="text-3xl font-extrabold mt-0.5">{sku}</div>
      <div class="text-slate-400 text-sm">{prod["name"] or ""}</div>
    </div>

    <div class="rounded-2xl bg-slate-950/40 border border-slate-700/50 p-4 text-sm space-y-1.5 text-slate-300">
      <div class="flex justify-between"><span>Lead Time</span><b>{int(prod["lead_time_days"])} gün</b></div>
      <div class="flex justify-between"><span>Z değeri</span><b>{float(prod["z_value"])}</b></div>
      <div class="flex justify-between"><span>FBA stok</span><b>{int(prod["fba_stock"])}</b></div>
      <div class="flex justify-between"><span>Yoldaki</span><b>{int(prod["inbound_stock"])}</b></div>
    </div>

    <div class="rounded-2xl bg-slate-950/40 border border-slate-700/50 p-4 text-center">
      <div class="text-xs text-slate-400 uppercase tracking-wide mb-1">60 Günlük Sipariş</div>
      <div class="text-5xl font-black {oq_color}">{oq}</div>
    </div>

    <a href="/?prefill={sku}" class="block text-center w-full rounded-2xl bg-indigo-500/15 text-indigo-300 border border-indigo-500/30 hover:bg-indigo-500/25 transition px-4 py-2.5 text-sm font-semibold">
      ✏️ Ürünü Düzenle
    </a>
    <form method="post" action="/delete/product/{sku}"
          onsubmit="return confirm('{sku} ve TÜM satış verileri kalıcı olarak silinsin mi?');">
      <button type="submit" class="w-full rounded-2xl bg-rose-500/10 text-rose-300 border border-rose-500/30 hover:bg-rose-500/20 transition px-4 py-2.5 text-sm font-semibold">
        🗑 SKU ve Satışları Sil
      </button>
    </form>
  </div>

  <!-- Sağ panel -->
  <div class="lg:col-span-2 space-y-6">
    <!-- Metrikler -->
    <div class="card">
      <div class="text-xs text-slate-400 uppercase tracking-wide mb-4">Son 3 Ay Hesabı</div>
      <div class="text-slate-300 text-xs mb-4">
        Baz alınan aylar: <b>{", ".join(month_label(y,m) for y,m in last3)}</b>
        &nbsp;|&nbsp; Bulunan veri: <b>{monthly_found if monthly_found else "—"}</b>
      </div>
      <div class="grid grid-cols-2 md:grid-cols-4 gap-4">
        <div class="rounded-2xl bg-slate-950/40 border border-slate-700/50 p-4">
          <div class="text-xs text-slate-400">Günlük Hız</div>
          <div class="text-2xl font-bold mt-1">{res["daily_velocity"]:.2f}</div>
        </div>
        <div class="rounded-2xl bg-slate-950/40 border border-slate-700/50 p-4">
          <div class="text-xs text-slate-400">Std Dev</div>
          <div class="text-2xl font-bold mt-1">{res["std_daily"]:.2f}</div>
        </div>
        <div class="rounded-2xl bg-slate-950/40 border border-slate-700/50 p-4">
          <div class="text-xs text-slate-400">Güvenlik Stoğu</div>
          <div class="text-2xl font-bold mt-1">{res["safety_stock"]:.2f}</div>
        </div>
        <div class="rounded-2xl bg-slate-950/40 border border-slate-700/50 p-4">
          <div class="text-xs text-slate-400">ROP</div>
          <div class="text-2xl font-bold mt-1">{res["rop"]:.2f}</div>
        </div>
      </div>
    </div>

    <!-- Grafik -->
    <div class="card">
      <div class="text-xs text-slate-400 uppercase tracking-wide mb-4">Son 6 Ay Satış Trendi</div>
      <canvas id="salesChart" height="110"></canvas>
    </div>

    <!-- Tek ay sil -->
    <div class="card">
      <h3 class="font-semibold mb-3">Belirli Bir Ay Verisini Sil</h3>
      <form method="post" action="/delete/sale"
            onsubmit="return confirm('Seçili aya ait satış verisi silinsin mi?');"
            class="flex flex-wrap gap-3">
        <input type="hidden" name="sku" value="{sku}">
        <input class="rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-2.5 text-sm" name="year" type="number" placeholder="Yıl" required/>
        <input class="rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-2.5 text-sm w-24" name="month" type="number" min="1" max="12" placeholder="Ay" required/>
        <button type="submit" class="rounded-2xl bg-rose-500/10 text-rose-300 border border-rose-500/30 hover:bg-rose-500/20 transition px-5 py-2.5 text-sm font-semibold">
          Bu Ayı Sil
        </button>
      </form>
    </div>
  </div>
</div>

<script>
  new Chart(document.getElementById('salesChart').getContext('2d'), {{
    type: 'bar',
    data: {{
      labels: {json.dumps(labels6)},
      datasets: [{{
        label: 'Aylık Satış',
        data: {json.dumps(units6)},
        backgroundColor: 'rgba(99,102,241,0.5)',
        borderColor: 'rgba(99,102,241,1)',
        borderWidth: 2,
        borderRadius: 6,
      }}]
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ labels: {{ color: '#94a3b8' }} }} }},
      scales: {{
        x: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: '#1e293b' }} }},
        y: {{ ticks: {{ color: '#94a3b8' }}, grid: {{ color: '#1e293b' }}, beginAtZero: true }}
      }}
    }}
  }});
</script>"""
    return page_shell(f"Ürün Detay — {sku}", body)


@app.post("/delete/product/{sku}")
def delete_product(sku: str, session: Optional[str] = Cookie(default=None)):
    user_id = _require_user(session)
    conn = get_conn()
    cur = conn.cursor()
    delete_product_and_sales(cur, user_id, sku)
    conn.commit()
    conn.close()
    return RedirectResponse(url="/products", status_code=303)


@app.post("/delete/sale")
def delete_sale(
    session: Optional[str] = Cookie(default=None),
    sku: str = Form(...),
    year: int = Form(...),
    month: int = Form(...),
):
    user_id = _require_user(session)
    conn = get_conn()
    cur = conn.cursor()
    delete_single_month_sale(cur, user_id, sku, year, month)
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/product/{sku}", status_code=303)


# ══════════════════════════════════════════════════════════════════════════════
#  Excel export
# ══════════════════════════════════════════════════════════════════════════════

def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(45, max(10, width + 2))


def _stream_wb(wb: Workbook, filename: str) -> StreamingResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/export/products.xlsx")
def export_products_xlsx(session: Optional[str] = Cookie(default=None)):
    user_id = _require_user(session)
    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute(
        "SELECT sku FROM products WHERE user_id=? ORDER BY updated_at DESC", (user_id,)
    ).fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "SKUListesi"
    ws.append(["SKU", "Ürün", "Lead Time", "Z", "FBA", "Yoldaki",
               "Günlük Hız", "Std Dev", "Güvenlik Stoğu", "ROP", "Sipariş (60g)"])
    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, user_id, sku)
        ws.append([
            sku, prod["name"] or "",
            int(prod["lead_time_days"]), float(prod["z_value"]),
            int(prod["fba_stock"]), int(prod["inbound_stock"]),
            round(res["daily_velocity"], 4), round(res["std_daily"], 4),
            round(res["safety_stock"], 4), round(res["rop"], 4),
            int(round(res["order_qty"])),
        ])
    conn.close()
    _autosize(ws)
    return _stream_wb(wb, "sku_listesi.xlsx")


@app.get("/export/plan.xlsx")
def export_plan_xlsx(session: Optional[str] = Cookie(default=None)):
    user_id = _require_user(session)
    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute(
        "SELECT sku FROM products WHERE user_id=? ORDER BY updated_at DESC", (user_id,)
    ).fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "SiparisListesi"
    ws.append(["SKU", "Ürün", "Lead Time", "Z", "FBA", "Yoldaki",
               "Günlük Hız", "ROP", "Sipariş (60g)"])
    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, user_id, sku)
        oq = int(round(res["order_qty"]))
        if oq <= 0:
            continue
        ws.append([
            sku, prod["name"] or "",
            int(prod["lead_time_days"]), float(prod["z_value"]),
            int(prod["fba_stock"]), int(prod["inbound_stock"]),
            round(res["daily_velocity"], 4), round(res["rop"], 4), oq,
        ])
    conn.close()
    _autosize(ws)
    return _stream_wb(wb, "siparis_listesi.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
#  Util
# ══════════════════════════════════════════════════════════════════════════════

def _get_user_or_redirect(session: Optional[str]):
    """Returns user_id int or RedirectResponse."""
    uid = _current_user_id(session)
    if uid is None:
        return RedirectResponse("/login", status_code=302)
    return uid
